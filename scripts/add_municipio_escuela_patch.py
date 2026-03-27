"""add_municipio_escuela_patch.py — Agrega MUNICIPIO_ESCUELA a data_dashboard.json

Lee escuelas.xlsx y construye un lookup {nombre_escuela → municipio}.
Luego inyecta el campo MUNICIPIO_ESCUELA en cada registro de data_dashboard.json
usando el campo ESCUELA como llave de join.

Corrección aplicada: "CORRIGIDORA" → "CORREGIDORA" en el campo Municipio del xlsx.

Uso:
    cd d:\\Metrix\\DashboardBecasLocal
    python scripts/add_municipio_escuela_patch.py
    python scripts/split_data.py
"""

import json
import pathlib
import sys
import openpyxl

ROOT      = pathlib.Path(__file__).resolve().parent.parent
XLSX_PATH = pathlib.Path(__file__).resolve().parent / "escuelas.xlsx"
DST_DASH  = ROOT / "data_dashboard.json"
TMP_PATH  = ROOT / "data_dashboard_mun_tmp.json"


def _norm(s: str) -> str:
    """Normaliza nombre para join: strip + upper + espacios colapsados."""
    return " ".join(str(s).upper().split()) if s else ""


def cargar_lookup_escuelas() -> dict:
    """Devuelve {nombre_normalizado: municipio} desde escuelas.xlsx."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_nombre    = header.index("Nombre")
        col_municipio = header.index("Municipio")
    except ValueError as e:
        sys.exit(f"ERROR: columna no encontrada en escuelas.xlsx: {e}")

    lookup: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre    = row[col_nombre]
        municipio = row[col_municipio]
        if not nombre:
            continue
        # Corrección de typo
        if isinstance(municipio, str):
            municipio = municipio.strip()
            if municipio.upper() == "CORRIGIDORA":
                municipio = "CORREGIDORA"
            else:
                municipio = municipio.upper().strip()
        lookup[_norm(nombre)] = municipio or ""

    wb.close()
    return lookup


def main() -> None:
    if not XLSX_PATH.exists():
        sys.exit(f"ERROR: no se encontró {XLSX_PATH}")
    if not DST_DASH.exists():
        sys.exit(f"ERROR: no se encontró {DST_DASH}")

    print("Cargando lookup desde escuelas.xlsx...")
    lookup = cargar_lookup_escuelas()
    print(f"  {len(lookup):,} escuelas en el lookup")

    sin_match = set()
    procesados = 0

    print(f"Inyectando MUNICIPIO_ESCUELA en {DST_DASH}...")
    with DST_DASH.open("r", encoding="utf-8") as fin, \
         TMP_PATH.open("w", encoding="utf-8") as fout:

        for raw in fin:
            raw = raw.strip()
            if not raw:
                continue
            rec = json.loads(raw)

            escuela = _norm(rec.get("ESCUELA", ""))
            mun_esc = lookup.get(escuela, "")
            if not mun_esc and escuela:
                sin_match.add(escuela)

            rec["MUNICIPIO_ESCUELA"] = mun_esc
            fout.write(json.dumps(rec, ensure_ascii=False) + "\n")
            procesados += 1

            if procesados % 10_000 == 0:
                print(f"  {procesados:,} registros procesados...", end="\r", flush=True)

    # Reemplazar archivo original
    TMP_PATH.replace(DST_DASH)

    print(f"\nListo: {procesados:,} registros actualizados")
    print(f"  Sin match en xlsx: {len(sin_match):,} escuelas únicas")
    if sin_match:
        print("  Primeras sin match:")
        for s in sorted(sin_match)[:20]:
            print(f"    · {s}")


if __name__ == "__main__":
    main()
